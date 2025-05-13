from math import log
import pandas as pd
import os
from openpyxl import load_workbook
import webbrowser
from openpyxl.styles import PatternFill, Font
#Función de impresión de los valores del Interes Compuesto
def datos_Icompuesto():
    print("""
    Bienvenido estás son las variables del Interés Compuesto:
    F = Monto o Valor Futuro
    P = Capital inicial o inversión inicial
    r = Tasa de interés
    t = Tiempo transcurrido o plazo
    """)   

def obtener_factor_de_conversion(conversion):
    disponible = False
    # Factores de conversión de cada formato a una base anual
    conversiones = {
        "anual": 1,
        "semestral": 2,
        "cuatrimestral": 3,
        "trimestral": 4,
        "bimestral": 6,
        "mensual": 12,
        "quincenal": 24,
        "semanal": 52,
        "diaria": 365
    }
    if conversion in conversiones:
        disponible = True
    return (disponible, conversiones[conversion])

#Funciones de conversión de los tiempos en las tasas y el plazo de tiempo
def conversor_tasa_c (tasa_interés_compuesta):
    try:
        # Extraer y convertir la tasa de interés
        command = tasa_interés_compuesta.split(" ")
        tiempo = float(command[0].rstrip("%")) / 100
        formato = command[1].lower()
        capitalizacion = command[3].lower()

        disponible, formato = obtener_factor_de_conversion(formato)
        if disponible == False:
            return tiempo
        
        disponible, capitalizacion = obtener_factor_de_conversion(capitalizacion)
        if disponible == False:
            return tiempo
        
        return tiempo * formato/capitalizacion

    except (IndexError, ValueError):
        return float(tasa_interés_compuesta[0])



def conversor_nplazos (command):
    plazo = command[0].split(" ")
    cap = command[1]
    try:

        tiempo = float(plazo[0])
        formato = (plazo[1]).lower()
        disponible, cap = obtener_factor_de_conversion(cap)
        if formato in ["dias","dia","día"]:
            n = ( tiempo / 30.44 ) * (12 / cap)
        elif formato in ["semana", "semanas"]:
            n = ( tiempo / 4.3 ) * (12 / cap)
        elif formato in ["quincena", "quincenas"]:
            n = ( tiempo * 0.5 ) * (12 / cap)
        elif formato in ["mes", "meses"]:
            n = ( tiempo ) * (12 / cap)
        elif formato in ["bimestre", "bimestres"]:
            n = ( tiempo * 2 ) * (12 / cap)
        elif formato in ["trimestre", "trimestres"]:
            n = ( tiempo * 3 ) * (12 / cap)
        elif formato in ["cuatrimestre", "cuatrimestres"]:
            n = ( tiempo * 4 ) * (12 / cap)
        elif formato in ["semestres", "semestre"]:
            n = ( tiempo * 6 ) * (12 / cap)
        elif formato in ["años","año"]:
            n = ( tiempo * 12 ) * (12 / cap)
        elif formato in ["decada","década","décadas","decadas"]:
            n = ( tiempo ) * (12 / cap)
        else:
            n = tiempo * (12 / cap)
        return (n)
    except (IndexError,ValueError):
        return(float(plazo[0]))
    
#Funciones para pedir los valores 
def Capital():
    P = float(input("Introduzca el Capital inicial o inversión inicial \n"))
    return (P)
def tasa_interés_compuesta():
    tasa_de_interés_compuesta = input("Introduzca la tasa de interés (indica si es diaria, semanal, mensual, bimestral, trimestral, cuatrimestral, semestral o anual): "
    "Ejemplo: 15% anual compuesto mensual \n")
    return (tasa_de_interés_compuesta)
def plazo():
    plazo = input("Introduzca el plazo (indica si es en: días, semanas, quincenas, meses, bimestres, trimestres, cuatrimestres, semestres, años o décadas): "
    "Ejemplo: 5 meses \n")
    cap = input("Porfavor si puedes introducir el tipo de capitalizacion que es la tasa:(anual, semestral, cuatrimestral, trimestral, bimestral, mensual, quincenal, diaria) \n")
    return (plazo,cap)
def Monto():
    F = float(input("Introduzca el Monto: \n"))
    return(F)
#Funciones para eficientar a la hora de pedir los valores
def IntereC_F ():
    P = Capital()
    r = conversor_tasa_c(tasa_interés_compuesta())
    t = conversor_nplazos(plazo())
    return(P,r,t)

def IntereC_P():
    F = Monto()
    r = conversor_tasa_c(tasa_interés_compuesta())
    t = conversor_nplazos(plazo())
    return(F,r,t)

def IntereC_r():
    t = conversor_nplazos(plazo())
    F = Monto()
    P = Capital()
    return(t,F,P)

def IntereC_t():
    F = Monto ()
    r = conversor_tasa_c(tasa_interés_compuesta())
    P = Capital()
    return(F,r,P)

#Funciones para hacer los procedimientos con las fórmulas
def Ic_monto (P,r,t):
    F = P * ( ( 1 + r ) ** t )
    return(F)

def Ic_capital(F,r,t):
    P = F / ( ( 1 + r ) ** t )  
    return(P)

def Ic_r(F,P,t):
    r = ( (F/P)**(1/t) ) -1
    return(r)

def Ic_tiempo(F,P,r):
    t = ( log (F/P) ) / (log (1+r) )
    return(t)

#Crea el archivo de Excel si no existe
if not os.path.exists("Análisis de Datos.xlsx"):
    pd.DataFrame(columns=["Capital","Tasa de Interés","Tiempo","Monto"]).to_excel("Interes_Compuesto.xlsx", index=False)

#Función para guardar los resultados en un archivo de Excel
from openpyxl import Workbook

from openpyxl.styles import PatternFill, Font

def guardar_en_analisis(tipo_interes, resultado_data, columna_resaltar):
    archivo = "Análisis de Datos.xlsx"
    
    try:
        # Crear el archivo si no existe
        if not os.path.exists(archivo):
            wb = Workbook()
            # Crear hojas         
            ws_compuesto = wb.create_sheet("Interés Compuesto")
            ws_compuesto.append(["Capital", "Tasa de Interés", "Tiempo", "Monto"])
            
            ws_simple = wb.create_sheet("Interés Simple")
            ws_simple.append(["Capital", "Tasa de Interés", "Tiempo", "Monto", "Intereses"])
            
            wb.save(archivo)
        
        # Cargar el archivo existente
        wb = load_workbook(archivo)
        
        # Agregar datos a la hoja específica
        if tipo_interes == "Compuesto":
            ws_compuesto = wb["Interés Compuesto"]
            fila_compuesto = [
                resultado_data.get("Capital", ""),
                resultado_data.get("Tasa de Interés", ""),
                resultado_data.get("Tiempo", ""),
                resultado_data.get("Monto", "")
            ]
            ws_compuesto.append(fila_compuesto)
            
            # Resaltar el resultado en la hoja "Interés Compuesto"
            if columna_resaltar in ["Capital", "Tasa de Interés", "Tiempo", "Monto"]:
                columna_index = ["Capital", "Tasa de Interés", "Tiempo", "Monto"].index(columna_resaltar) + 1
                ultima_fila_compuesto = ws_compuesto.max_row
                ws_compuesto.cell(row=ultima_fila_compuesto, column=columna_index).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws_compuesto.cell(row=ultima_fila_compuesto, column=columna_index).font = Font(bold=True)
        
        elif tipo_interes == "Simple":
            ws_simple = wb["Interés Simple"]
            fila_simple = [
                resultado_data.get("Capital", ""),
                resultado_data.get("Tasa de Interés", ""),
                resultado_data.get("Tiempo", ""),
                resultado_data.get("Monto", ""),
                resultado_data.get("Intereses", "")
            ]
            ws_simple.append(fila_simple)
            
            # Resaltar el resultado en la hoja "Interés Simple"
            if columna_resaltar in ["Capital", "Tasa de Interés", "Tiempo", "Monto", "Intereses"]:
                columna_index = ["Capital", "Tasa de Interés", "Tiempo", "Monto", "Intereses"].index(columna_resaltar) + 1
                ultima_fila_simple = ws_simple.max_row
                ws_simple.cell(row=ultima_fila_simple, column=columna_index).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws_simple.cell(row=ultima_fila_simple, column=columna_index).font = Font(bold=True)
        
        # Guardar los cambios
        wb.save(archivo)
        print(f"Los resultados han sido guardados en el archivo {archivo}")
    
    except PermissionError:
        print(f"Error: No se puede acceder al archivo '{archivo}'. Asegúrate de que no esté abierto en otra aplicación.")
    except Exception as e:
        print(f"Error inesperado al guardar los datos: {e}")
        
#Función para sacar el valor pedido y el resultado
def Opciones ():
    opcion = input("Bienvenido al Interes Compuesto, ¿Qué deseas calcular? (F, P, r, t) // Si desea terminar el programa solo escriba salir: ") 
    Valor_pedido = None
    Resultado = None
    if opcion.lower() == "salir":
        Resultado = ""
        Valor_pedido = "salir"
    elif opcion.lower() != "salir":      
        if opcion.upper() == "F":
            P,r,t = IntereC_F ()
            Valor_pedido = ">>> El monto es de"
            Resultado = Ic_monto (P,r,t)
            #Valores de guardado
            resultado_data = {"Capital": P, "Tasa de Interés": r, "Tiempo": t, "Monto": Resultado}
            columna_resaltar = "Monto"
            guardar_en_analisis("Compuesto", resultado_data, columna_resaltar)
        elif opcion.upper() == "P":
            F,r,t = IntereC_P () 
            Valor_pedido = ">>> El capital es de"
            Resultado = Ic_capital (F,r,t)
            #Valores de guardado
            resultado_data = {"Capital": Resultado, "Tasa de Interés": r, "Tiempo": t, "Monto": F}
            columna_resaltar = "Capital"
            guardar_en_analisis("Compuesto", resultado_data, columna_resaltar)
        elif opcion.lower() == "r":
            t,F,P = IntereC_r()
            Valor_pedido = ">>> La tasa de interes es"
            Resultado = Ic_r(t,F,P)
            #Valores de guardado
            resultado_data = {"Capital": P, "Tasa de Interés": Resultado, "Tiempo": t, "Monto": F}
            columna_resaltar = "Tasa de Interés"  
            guardar_en_analisis("Compuesto", resultado_data, columna_resaltar)  
        elif opcion.lower() == "t":
            F,r,P = IntereC_t ()
            Valor_pedido = ">>> El plazo de tiempo es de"
            Resultado = Ic_tiempo(F,P,r)
            #Valores de guardado
            resultado_data = {"Capital": P, "Tasa de Interés": r, "Tiempo": Resultado, "Monto": F}
            columna_resaltar = "Tiempo"
            guardar_en_analisis("Compuesto", resultado_data, columna_resaltar)
    else:   
        Resultado = print ("Porfavor imprime Números validos")
        Valor_pedido = ""
    return(Valor_pedido,Resultado)

#Función Principal del Interes Compuesto
def Interes_Compuesto ():
    datos_Icompuesto ()
    contador = 0
    while True:
        Valor_Pedido, Resultado = Opciones()
        if Valor_Pedido == "salir":
            break
        elif Valor_Pedido == "":
            print(Resultado)
        else: 
            print(f"{Valor_Pedido}: {Resultado}")
    print("El programa ha terminado")